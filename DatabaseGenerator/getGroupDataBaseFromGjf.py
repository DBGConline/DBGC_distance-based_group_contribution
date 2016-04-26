#---------------------------------------------------------------------------------------------------
# This code is used for assist the use of distance-based group contribution (DBGC) method.
# The code is published under the MIT open source license.

# The MIT License (MIT)
# Copyright (c) 2016 by the DBGC Team (DBGConline@gmail.com)

# Permission is hereby granted, free of charge, to any person obtaining a copy of this software 
# and associated documentation files (the "Software"), to deal in the Software without restriction, 
# including without limitation the rights to use, copy, modify, merge, publish, distribute, 
# sublicense, and/or sell copies of the Software, and to permit persons to whom the Software is 
# furnished to do so, subject to the following conditions:

# The above copyright notice and this permission notice shall be included in all copies or 
# substantial portions of the Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING 
# BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND 
# NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, 
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, 
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.
#---------------------------------------------------------------------------------------------------

'''
This file is used to generate the group contribution veoctors for the species to be predicted.

By default the input file is stored in the directiry Gjfs. The input format is the Gaussian input 
format, namely, .gjf. Here is an example.

	-----------------example below--------------------------------
	# hf/3-21g 

	Title Card Required

	0 1
	 C                 -0.68027212   -0.34013605    0.00000000
	 H                 -0.32361770   -1.34894605    0.00000000
	 H                 -0.32359928    0.16426214   -0.87365150
	 H                 -1.75027212   -0.34012287    0.00000000
	 C                 -0.16692990    0.38582022    1.25740497
	 H                 -0.52198625    1.39519331    1.25642745
	 H                 -0.52519901   -0.11744821    2.13105486
	 C                  1.37306749    0.38336387    1.25881364
	 H                  1.73133598    0.88630575    0.38497547
	 H                  1.72974052    0.88808796    2.13227684
	 H                  1.72812373   -0.62600882    1.26016738



	 ------------------example above-------------------------------
The first line is the command line for Gaussian calculation, so it doesn't matter. The information 
needed to generate group contribution vector is the coordinates.

The output file is DBGCVectors.xlsx and DBGCVectors_noTemplate.xlsx in the directory DBGCVectors, which 
contains the group contribution vectors. In a system with finite groups, the dimension of the group 
contribution vector is a constant. A template file is used when the number of species is too small to 
cover all the groups. Currently, the template file used is groupTemplate.xlsx, containing all the groups 
in alkane, alkene (only one C=C double bond is considered), alkyl radical, and alkenyl radical system. If 
the types of groups are not determined, DBGCVectors_noTemplate.xlsx are used to record all the groups and 
group-group interactions in the new system, which consists of the .gjf files in the directory Gjfs.

'''

import os
import openpyxl

import groupCounter

# if using the group Template
groupCounter.groupCounter.readGroupTemplate()

counterA = groupCounter.groupCounter()

tmp_fileList = os.listdir('Gjfs')
for tmp_file in tmp_fileList:
	counterA.readGjfFile(fileName=tmp_file, directory='Gjfs', moleculeLabel=tmp_file[0:-4])
	counterA.writeDBGCVector(overwrite=False)

# if not using the group Template
# definition of variables
all_moles = []
all_groups = set()
groupIndex = {}
fileName='DBGCVectors_noTemplate.xlsx'
speciesNumber = 0

# definition of temporary variables
tmp_groups= []

for tmp_file in tmp_fileList:
	counterA = groupCounter.groupCounter()
	counterA.readGjfFile(fileName=tmp_file, directory='Gjfs', moleculeLabel=tmp_file[0:-4])
	tmp_groups = counterA.mole.get1stOrderGroup()
	all_moles.append(counterA.mole)
	all_groups = all_groups | set(tmp_groups)
all_groups = sorted(list(all_groups))
N = len(all_groups)
n = len(all_moles)


if not os.path.exists('DBGCVectors'):
	os.mkdir('DBGCVectors')
os.chdir('DBGCVectors')

vectorDimension=N*(N+3)/2
wbw = openpyxl.Workbook()
shw = wbw.active
shw.title = 'inputVectors'
tmp_row = 1
tmp_col = 1
shw.cell(row=tmp_row, column=tmp_col).value = 'ID'
shw.cell(row=tmp_row, column=tmp_col+1).value = 'Number of species'
shw.cell(row=tmp_row, column=tmp_col+3).value = 'TotalDimesionNumber'
shw.cell(row=tmp_row, column=tmp_col+5).value = 'DimensionIndex'
tmp_row = 2
shw.cell(row=tmp_row, column=tmp_col+3).value = vectorDimension		
tmp_col = 6
for i in xrange(N):
	shw.cell(row=tmp_row, column=tmp_col).value = tmp_col-5
	shw.cell(row=tmp_row+1, column=tmp_col).value = all_groups[i]
	groupIndex[all_groups[i]] = tmp_col
	tmp_col += 1
for i in xrange(N):
	for j in xrange(i, N):
		tmp_list = sorted([all_groups[i], all_groups[j]])
		tmp_text = tmp_list[0] + '-' + tmp_list[1]
		shw.cell(row=tmp_row, column=tmp_col).value = tmp_col-5
		shw.cell(row=tmp_row+1, column=tmp_col).value = tmp_text
		groupIndex[tmp_text] = tmp_col
		tmp_col += 1
tmp_row = 3
shw.cell(row=tmp_row, column=vectorDimension+8).value = 'Name'
shw.cell(row=tmp_row, column=vectorDimension+9).value = 'ReferenceEnergy'

for tmp_mole in all_moles:			
	tmp_groupVector = tmp_mole.getGroupVector()
	tmp_row = 4+speciesNumber
	tmp_col = 1
	speciesNumber += 1
	shw.cell(row=tmp_row, column=tmp_col).value = speciesNumber
	for i in xrange(vectorDimension):
		shw.cell(row=tmp_row, column=tmp_col+5+i).value = 0.0
	for tmp_vectorEle in tmp_groupVector.keys():
		shw.cell(row=tmp_row, column=groupIndex[tmp_vectorEle]).value = tmp_groupVector[tmp_vectorEle]
	shw.cell(row=tmp_row, column=vectorDimension+8).value = tmp_mole.label

	tmp_row = 2
	tmp_col = 2
	shw.cell(row=tmp_row, column=tmp_col).value = speciesNumber

wbw.save(fileName)
os.chdir('../')

